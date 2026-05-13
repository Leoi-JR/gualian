"""
关键词规则验证器 - Keyword Rule Validator

功能描述：
这个模块用于验证Excel文件中关键词的格式是否符合预定义的规则模式。
主要应用场景包括内容过滤、关键词搜索、数据质量检查等。

核心功能：
1. 支持10种不同的关键词组合规则模式
2. 自动处理中文符号转换
3. 批量验证Excel文件中的关键词列
4. 生成错误报告和高亮标记文件

作者：系统自动整理
日期：2024年
"""

import re
import pandas as pd
import numpy as np
from tqdm import tqdm
from datetime import datetime
import sys
from pathlib import Path

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from config.manager import config_manager
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class KeywordRuleValidator:
    """关键词规则验证器类"""
    
    def __init__(self):
        """初始化验证器，加载配置和预编译正则表达式"""
        # 从JSON配置文件加载设置
        self.excel_file_path = config_manager.get_str('excel.file_path')
        self.excel_sheet_name = config_manager.get_str('excel.sheet_name')
        self.keyword_columns = config_manager.get_list('keyword_columns')

        # 列映射配置
        column_mappings = config_manager.get_dict('column_mappings')
        self.like_keyword_column = column_mappings.get('like_keyword_column', 'like_keyword')
        self.must_keyword_column = column_mappings.get('must_keyword_column', 'must_keyword')
        self.unlike_keyword_column = column_mappings.get('unlike_keyword_column', 'unlike_keyword')

        # 验证配置
        validation_rules = config_manager.get_dict('validation_rules')
        self.highlight_color = validation_rules.get('highlight_color', 'FFFF00')
        self.max_errors_display = validation_rules.get('max_errors_display', 10)

        # 新增字段验证配置
        matching_rules = config_manager.get_dict('keyword_matching.matching_rules')

        # type_filter字段配置
        self.industry_filter_config = matching_rules.get('industry_filter', {})
        self.industry_separator = self.industry_filter_config.get('separator', '、')
        self.industry_default_value = self.industry_filter_config.get('default_value', 'default')

        # field_scope字段配置
        self.source_scope_config = matching_rules.get('source_scope_filter', {})
        self.source_scope_separator = self.source_scope_config.get('separator', '、')
        self.source_scope_default_value = self.source_scope_config.get('default_value', 'default')

        # 获取field_scope的有效值列表
        except_mappings = self.source_scope_config.get('except_mappings', {})
        self.valid_source_scope_values = []
        for mapping_values in except_mappings.keys():
            self.valid_source_scope_values.append(mapping_values)

        # 添加需要检查的新字段
        self.additional_check_columns = ['type_filter', 'field_scope']
        
        # 预编译正则表达式模式
        self.compiled_patterns = [
            (re.compile(r'^([^,(){}&]+)&([^,(){}&]+)$'), 1, "A&B"),                          
            (re.compile(r'^([^,(){}&]+)&&([^,(){}&]+)$'), 2, "A&&B"),                        
            (re.compile(r'^([^,(){}&]+)&\(([^,(){}&]+(?:,[^,(){}&]+)*)\)$'), 3, "A&(B,C,D)"),       
            (re.compile(r'^\(([^,(){}&]+(?:,[^,(){}&]+)*)\)&([^,(){}&]+)$'), 4, "(B,C,D)&A"),       
            (re.compile(r'^([^,(){}&]+)&&\(([^,(){}&]+(?:,[^,(){}&]+)*)\)$'), 5, "A&&(B,C,D)"),     
            (re.compile(r'^\(([^,(){}&]+(?:,[^,(){}&]+)*)\)&&([^,(){}&]+)$'), 6, "(B,C,D)&&A"),     
            (re.compile(r'^\(([^,(){}&]+(?:,[^,(){}&]+)*)\)&\(([^,(){}&]+(?:,[^,(){}&]+)*)\)$'), 7, "(A,B)&(C,D)"),  
            (re.compile(r'^\(([^,(){}&]+(?:,[^,(){}&]+)*)\)&&\(([^,(){}&]+(?:,[^,(){}&]+)*)\)$'), 8, "(A,B)&&(C,D)"), 
            (re.compile(r'^([^,(){}&]+)\.\{0,(\d+)\}([^,(){}&]+)$'), 9, "A.{0,X}B"),              
            (re.compile(r'^([^,(){}&]+)$'), 10, "A")                                     
        ]
    
    def check_rule_pattern(self, text):
        """
        检查文本是否匹配任何规则模式

        Args:
            text (str): 要检查的文本

        Returns:
            int|bool: 匹配的规则编号，或False（不匹配）
        """
        # 去除空格
        text = text.strip()

        # 清理不可见字符（零宽度空格等）
        text = self.clean_invisible_characters(text)

        # 空值处理
        if text == '0' or text == '':
            return True

        # 使用预编译的正则表达式检查文本
        for pattern, rule_number, description in self.compiled_patterns:
            match = pattern.match(text)
            if match:
                return rule_number

        return False

    def clean_invisible_characters(self, text):
        """
        清理文本中的不可见字符

        Args:
            text (str): 输入文本

        Returns:
            str: 清理后的文本
        """
        if not text:
            return text

        # 常见的不可见字符Unicode码点
        invisible_chars = [
            '\u200B',  # 零宽度空格 (Zero Width Space)
            '\u200C',  # 零宽度非连接符 (Zero Width Non-Joiner)
            '\u200D',  # 零宽度连接符 (Zero Width Joiner)
            '\u2060',  # 词连接符 (Word Joiner)
            '\uFEFF',  # 零宽度非断空格 (Zero Width No-Break Space)
            '\u00A0',  # 非断空格 (Non-Breaking Space)
        ]

        # 移除这些不可见字符
        for char in invisible_chars:
            text = text.replace(char, '')

        return text
    
    def replace_chinese_symbols(self, text):
        """
        将中文符号替换为英文符号
        
        Args:
            text: 输入文本
            
        Returns:
            str: 替换后的文本
        """
        if pd.notna(text):
            # 中文符号转英文符号
            text = str(text).replace('，', ',')
            text = text.replace('（', '(')
            text = text.replace('）', ')')
            # 去除空格
            text = text.replace(' ', '')
        return text

    def check_rule_value(self, value):
        """
        检查单元格的值是否符合规则（支持|分隔的多个值）

        Args:
            value: 单元格的值

        Returns:
            bool: 是否符合规则
        """
        result = self.check_rule_value_detailed(value)
        return result['valid']

    def check_rule_value_detailed(self, value):
        """
        检查单元格的值是否符合规则，返回详细的验证结果

        Args:
            value: 单元格的值

        Returns:
            dict: 包含验证结果和错误详情的字典
                {
                    'valid': bool,
                    'error_type': str,
                    'error_message': str
                }
        """
        if pd.isna(value):
            return {'valid': True, 'error_type': None, 'error_message': None}

        value_str = str(value).strip()

        # 空字符串或'0'被视为有效
        if value_str == '' or value_str == '0':
            return {'valid': True, 'error_type': None, 'error_message': None}

        # 检查是否包含连续的"|"符号
        if '||' in value_str:
            return {
                'valid': False,
                'error_type': 'consecutive_pipes',
                'error_message': '包含连续的"|"符号'
            }

        # 检查是否以"|"开头或结尾
        if value_str.startswith('|'):
            return {
                'valid': False,
                'error_type': 'starts_with_pipe',
                'error_message': '以"|"符号开头'
            }
        if value_str.endswith('|'):
            return {
                'valid': False,
                'error_type': 'ends_with_pipe',
                'error_message': '以"|"符号结尾'
            }

        # 按|分割多个值
        values = value_str.split('|')
        for i, v in enumerate(values):
            # 检查是否有空的分段
            if v.strip() == '':
                return {
                    'valid': False,
                    'error_type': 'empty_segment',
                    'error_message': f'第{i+1}个分段为空'
                }
            # 检查每个分段是否符合规则模式
            if not self.check_rule_pattern(v.strip()):
                return {
                    'valid': False,
                    'error_type': 'invalid_pattern',
                    'error_message': f'第{i+1}个分段"{v.strip()}"不符合关键词规则模式'
                }

        return {'valid': True, 'error_type': None, 'error_message': None}

    def check_industry_type_value(self, value):
        """
        检查industry_type字段的值是否符合规则

        Args:
            value: 单元格的值

        Returns:
            bool: 是否符合规则
        """
        result = self.check_industry_type_value_detailed(value)
        return result['valid']

    def check_industry_type_value_detailed(self, value):
        """
        检查industry_type字段的值是否符合规则，返回详细的验证结果

        Args:
            value: 单元格的值

        Returns:
            dict: 包含验证结果和错误详情的字典
        """
        if pd.isna(value):
            return {'valid': True, 'error_type': None, 'error_message': None}

        value_str = str(value).strip()

        # 空值或默认值
        if value_str == '' or value_str == self.industry_default_value:
            return {'valid': True, 'error_type': None, 'error_message': None}

        # 按分隔符分割
        parts = value_str.split(self.industry_separator)

        # 检查每个部分是否为单个大写字母
        for i, part in enumerate(parts):
            part = part.strip()
            if not part.isalpha():
                return {
                    'valid': False,
                    'error_type': 'not_alphabetic',
                    'error_message': f'第{i+1}个部分"{part}"不是字母'
                }
            elif len(part) != 1:
                return {
                    'valid': False,
                    'error_type': 'not_single_letter',
                    'error_message': f'第{i+1}个部分"{part}"不是单个字母'
                }
            elif not part.isupper():
                return {
                    'valid': False,
                    'error_type': 'not_uppercase',
                    'error_message': f'第{i+1}个部分"{part}"不是大写字母'
                }

        return {'valid': True, 'error_type': None, 'error_message': None}

    def check_source_scope_value(self, value):
        """
        检查source_scope字段的值是否符合规则

        Args:
            value: 单元格的值

        Returns:
            bool: 是否符合规则
        """
        result = self.check_source_scope_value_detailed(value)
        return result['valid']

    def check_source_scope_value_detailed(self, value):
        """
        检查source_scope字段的值是否符合规则，返回详细的验证结果

        Args:
            value: 单元格的值

        Returns:
            dict: 包含验证结果和错误详情的字典
        """
        if pd.isna(value):
            return {'valid': True, 'error_type': None, 'error_message': None}

        value_str = str(value).strip()

        # 空值或默认值
        if value_str == '' or value_str == self.source_scope_default_value:
            return {'valid': True, 'error_type': None, 'error_message': None}

        # 按分隔符分割
        parts = value_str.split(self.source_scope_separator)

        # 检查每个部分是否在有效值列表中
        for i, part in enumerate(parts):
            part = part.strip()
            if part not in self.valid_source_scope_values:
                return {
                    'valid': False,
                    'error_type': 'invalid_value',
                    'error_message': f'第{i+1}个部分"{part}"不在有效值列表中: {self.valid_source_scope_values}'
                }

        return {'valid': True, 'error_type': None, 'error_message': None}

    def highlight_error_cells(self, error_cells, df):
        """
        高亮错误单元格并保存新文件
        
        Args:
            error_cells (list): 错误单元格列表
            df (DataFrame): 数据框
            
        Returns:
            str: 新文件路径
        """
        # 加载工作簿
        wb = load_workbook(self.excel_file_path)
        ws = wb[self.excel_sheet_name]
        
        # 创建错误高亮填充样式
        highlight_fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type='solid')
        
        # 获取列名到列索引的映射
        col_indices = {col_name: idx+1 for idx, col_name in enumerate(df.columns)}
        
        # 高亮错误单元格
        for idx, col, value in error_cells:
            # pandas的行索引从0开始，Excel的行索引从1开始，且有表头，所以+2
            row_idx = idx + 2
            col_idx = col_indices[col]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = highlight_fill
            
            # 更新单元格的值为替换后的值
            cell.value = df.at[idx, col]
        
        # 生成带时间戳的新文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name, file_ext = os.path.splitext(self.excel_file_path)
        new_file_path = f"{file_name}_checked_highlighted_{timestamp}{file_ext}"
        
        # 保存工作簿
        wb.save(new_file_path)
        return new_file_path

    def has_data_changed(self, original_df, cleaned_df):
        """
        检查数据是否在清理过程中发生了变化

        Args:
            original_df (DataFrame): 原始数据
            cleaned_df (DataFrame): 清理后的数据

        Returns:
            bool: 数据是否有变化
        """
        # 检查关键词列是否有变化
        for col in self.keyword_columns:
            if col in original_df.columns and col in cleaned_df.columns:
                # 比较每个单元格的值
                for idx in original_df.index:
                    original_value = original_df.loc[idx, col]
                    cleaned_value = cleaned_df.loc[idx, col]

                    # 处理NaN值
                    if pd.isna(original_value) and pd.isna(cleaned_value):
                        continue
                    elif pd.isna(original_value) or pd.isna(cleaned_value):
                        return True
                    elif str(original_value) != str(cleaned_value):
                        return True

        return False

    def save_cleaned_file(self, df):
        """
        保存清理后的文件（无错误情况）

        Args:
            df (DataFrame): 清理后的数据

        Returns:
            str: 新文件路径
        """
        # 生成新文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = Path(self.excel_file_path).stem
        extension = Path(self.excel_file_path).suffix
        new_file_name = f"{base_name}_cleaned_{timestamp}{extension}"
        new_file_path = Path(self.excel_file_path).parent / new_file_name

        # 保存文件
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=self.excel_sheet_name, index=False)

        return str(new_file_path)

    def validate_excel_file(self):
        """
        验证Excel文件中的关键词规则

        Returns:
            dict: 验证结果
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(
                self.excel_file_path,
                sheet_name=self.excel_sheet_name,
                dtype={
                    self.like_keyword_column: str,
                    self.must_keyword_column: str,
                    self.unlike_keyword_column: str
                }
            )

            # 检查是否包含所需的列
            missing_columns = []
            for col in self.keyword_columns:
                if col not in df.columns:
                    missing_columns.append(col)

            if missing_columns:
                error_msg = f"错误：Excel文件中缺少以下列：{', '.join(missing_columns)}"
                raise ValueError(error_msg)

            # 保存原始数据副本用于比较
            original_df = df.copy()

            # 对关键词列进行中文符号替换
            for col in self.keyword_columns:
                df[col] = df[col].apply(self.replace_chinese_symbols)

            # 检查关键词列的规则，记录不符合规则的单元格
            error_cells = []
            total_cells_checked = 0

            # 检查关键词字段
            for col in self.keyword_columns:
                for idx, value in df[col].items():
                    total_cells_checked += 1
                    validation_result = self.check_rule_value_detailed(value)
                    if not validation_result['valid']:
                        error_cells.append((idx, col, value))
                        print(f"关键词规则错误 - 行{idx+2}, 列'{col}': '{value}' - {validation_result['error_message']}")

            # 检查type_filter字段（如果存在）
            if 'type_filter' in df.columns:
                for idx, value in df['type_filter'].items():
                    total_cells_checked += 1
                    validation_result = self.check_industry_type_value_detailed(value)
                    if not validation_result['valid']:
                        error_cells.append((idx, 'type_filter', value))
                        print(f"type_filter格式错误 - 行{idx+2}: '{value}' - {validation_result['error_message']} (应为'default'或大写字母用'、'连接，如'A、F、M')")

            # 检查field_scope字段（如果存在）
            if 'field_scope' in df.columns:
                for idx, value in df['field_scope'].items():
                    total_cells_checked += 1
                    validation_result = self.check_source_scope_value_detailed(value)
                    if not validation_result['valid']:
                        error_cells.append((idx, 'field_scope', value))
                        print(f"field_scope格式错误 - 行{idx+2}: '{value}' - {validation_result['error_message']} (应为'default'或有效值用'、'连接: {self.valid_source_scope_values})")

            # 检查数据是否有变化（中文符号替换、不可见字符清理等）
            data_changed = self.has_data_changed(original_df, df)

            # 准备返回结果
            result = {
                'success': True,
                'total_cells_checked': total_cells_checked,
                'error_count': len(error_cells),
                'error_cells': error_cells,
                'new_file_path': None,
                'data_cleaned': data_changed
            }

            # 如果有错误，高亮并保存新文件
            if error_cells:
                new_file_path = self.highlight_error_cells(error_cells, df)
                result['new_file_path'] = new_file_path
            # 如果没有错误但数据有变化，保存清理后的文件
            elif data_changed:
                new_file_path = self.save_cleaned_file(df)
                result['new_file_path'] = new_file_path

            return result

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'error_count': 0,
                'error_cells': []
            }

    def test_rule_patterns(self):
        """测试规则模式的示例"""
        test_cases = [
            ("keyword1&keyword2", 1, "简单AND连接"),
            ("keyword1&&keyword2", 2, "双AND连接"),
            ("keyword1&(keyword2,keyword3,keyword4)", 3, "单个与组合"),
            ("(keyword2,keyword3,keyword4)&keyword1", 4, "组合与单个"),
            ("keyword1&&(keyword2,keyword3,keyword4)", 5, "双AND与组合"),
            ("(keyword2,keyword3,keyword4)&&keyword1", 6, "组合与双AND"),
            ("(keyword1,keyword2)&(keyword3,keyword4)", 7, "组合与组合"),
            ("(keyword1,keyword2)&&(keyword3,keyword4)", 8, "双AND组合"),
            ("keyword1.{0,5}keyword2", 9, "距离限制"),
            ("keyword1", 10, "单个关键词"),
            ("keyword1&keyword2&keyword3", False, "不支持的格式"),
            ("keyword1&(keyword2", False, "语法错误")
        ]
        
        print("规则模式测试结果：")
        print("-" * 60)
        for test_case, expected, description in test_cases:
            result = self.check_rule_pattern(test_case)
            status = "✓" if result == expected else "✗"
            print(f"{status} '{test_case}' -> 规则{result} ({description})")
            
    def get_rule_descriptions(self):
        """获取所有规则的描述"""
        rules = [
            "规则1: A&B - 两个关键词用&连接",
            "规则2: A&&B - 两个关键词用&&连接",
            "规则3: A&(B,C,D) - 单个关键词与括号内多个关键词组合",
            "规则4: (B,C,D)&A - 括号内多个关键词与单个关键词组合",
            "规则5: A&&(B,C,D) - 使用&&的单个与组合",
            "规则6: (B,C,D)&&A - 使用&&的组合与单个",
            "规则7: (A,B)&(C,D) - 两组括号关键词的&组合",
            "规则8: (A,B)&&(C,D) - 两组括号关键词的&&组合",
            "规则9: A.{0,X}B - 正则表达式距离限制模式",
            "规则10: A - 单个关键词"
        ]
        return rules


def run_validation(debug_mode=False):
    """主函数 - 运行关键词规则验证

    Args:
        debug_mode (bool): 是否启用调试模式，显示详细信息
    """
    validator = KeywordRuleValidator()

    print("=" * 60)
    print("关键词规则验证器")
    print("=" * 60)

    # 调试模式下显示规则信息和测试结果
    if debug_mode:
        print("\n支持的规则模式：")
        for rule in validator.get_rule_descriptions():
            print(f"  {rule}")

        print("\n" + "=" * 40)
        validator.test_rule_patterns()
        print("\n" + "=" * 40)

    print("开始验证Excel文件...")

    result = validator.validate_excel_file()
    
    if result['success']:
        print(f"验证完成！")
        print(f"总共检查了 {result['total_cells_checked']} 个单元格")

        if result['error_count'] > 0:
            print(f"发现 {result['error_count']} 个不符合规则的单元格：")

            # 按字段类型分组显示错误
            keyword_errors = []
            industry_errors = []
            source_scope_errors = []

            for idx, col, value in result['error_cells']:
                if col in validator.keyword_columns:
                    keyword_errors.append((idx, col, value))
                elif col == 'type_filter':
                    industry_errors.append((idx, col, value))
                elif col == 'field_scope':
                    source_scope_errors.append((idx, col, value))

            if keyword_errors:
                print(f"\n关键词规则错误 ({len(keyword_errors)}个):")
                for idx, col, value in keyword_errors[:validator.max_errors_display]:
                    print(f"  行{idx+2}, 列'{col}': '{value}'")
                if len(keyword_errors) > validator.max_errors_display:
                    print(f"  ... 还有 {len(keyword_errors) - validator.max_errors_display} 个错误")

            if industry_errors:
                print(f"\ntype_filter字段错误 ({len(industry_errors)}个):")
                for idx, col, value in industry_errors[:validator.max_errors_display]:
                    print(f"  行{idx+2}: '{value}' (应为'default'或大写字母用'、'连接)")
                if len(industry_errors) > validator.max_errors_display:
                    print(f"  ... 还有 {len(industry_errors) - validator.max_errors_display} 个错误")

            if source_scope_errors:
                print(f"\nfield_scope字段错误 ({len(source_scope_errors)}个):")
                print(f"  有效值: {validator.valid_source_scope_values}")
                for idx, col, value in source_scope_errors[:validator.max_errors_display]:
                    print(f"  行{idx+2}: '{value}' (应为'default'或有效值用'、'连接)")
                if len(source_scope_errors) > validator.max_errors_display:
                    print(f"  ... 还有 {len(source_scope_errors) - validator.max_errors_display} 个错误")

            print(f"\n已生成高亮文件：{result['new_file_path']}")
        else:
            print("✓ 所有单元格都符合规则要求！")

            # 检查是否有数据清理
            if result.get('data_cleaned', False):
                print("📝 检测到数据清理操作（中文符号替换、不可见字符清理等）")
                print(f"已保存清理后的文件：{result['new_file_path']}")
            elif result.get('new_file_path'):
                print(f"已保存清理后的文件：{result['new_file_path']}")
    else:
        print(f"验证失败：{result['error']}")


def main():
    """主入口点，供其他模块调用"""
    # 检查是否启用调试模式
    debug_mode = os.getenv('KEYWORD_VALIDATOR_DEBUG', 'false').lower() == 'true'

    # 也可以通过命令行参数启用调试模式
    if len(sys.argv) > 1 and sys.argv[1] == '--debug':
        debug_mode = True

    run_validation(debug_mode)


if __name__ == "__main__":
    main()